import time
import subprocess
import pygetwindow as gw
import cv2
import pyautogui
import keyboard
import os
from pynput import mouse
from PIL import ImageGrab
import pyperclip
from openpyxl import load_workbook
from decimal import Decimal

"""获取剪贴板数据"""


def clipboard_get():
    time.sleep(0.2)
    data = pyperclip.paste()  # 主要这里差别
    return data


"""
# 查找输入框并输入字符串
"""

# def find_picture_click_and_input(orginal_path, target_path, original_name, target_name, input_word):
#     find_picture_click(orginal_path, target_path, original_name, target_name, "double")
#     # 输入字符串
#     pyautogui.typewrite( interval=0.1)
#     pyautogui.typewrite(input_word, interval=0.1)
#     keyboard.press_and_release("enter")


"""
# 查找图片并点击
# action : single: 单击 , double: 双击
"""


def find_picture_click(orginal_path, target_path, filename, action):
    # 找到最后一个圆点的索引位置
    last_dot_index = filename.rfind(".")
    if last_dot_index != -1:
        file_name = filename[:last_dot_index]  # 获取文件名部分
        file_extension = filename[last_dot_index + 1:]  # 获取文件名后缀部分
    else:
        print("无法解析文件名和文件名后缀")
    # 截图的保存路径的绝对路径名
    orginal_path_name = orginal_path + file_name + "_origin." + file_extension
    # 目标图片的保存路径的绝对路径名
    target_path_name = target_path + file_name + "." + file_extension

    # 截图
    pic = ImageGrab.grab()

    # 保存图片
    pic.save(orginal_path_name)

    # 获取目标图片左上角、右下角的坐标
    top_left, bottom_right = match_template(orginal_path_name, target_path_name)
    # print("左上角、右下角的坐标分别是：", (top_left, bottom_right))

    # 获取鼠标即将点击的x,y坐标
    x, y = get_click_point(top_left, bottom_right)
    # print("鼠标点击的[x,y]坐标是：", (x, y))

    if action == "single":
        # 单击鼠标左键
        # 参数1：移动x坐标到指定位置
        # 参数2：移动y坐标到指定位置
        # 参数3：点击次数
        # 无参数：在当前坐标单击
        pyautogui.click(x, y)
    elif action == "double":
        # 双击鼠标左键
        # 参数1：移动x坐标到指定位置
        # 参数2：移动y坐标到指定位置
        # 无参数：在当前坐标双击
        pyautogui.doubleClick(x, y)
    elif action == "right": 
        # 点击鼠标右键
        # 参数1：移动x坐标到指定位置
        # 参数2：移动y坐标到指定位置
        # 无参数：在当前坐标点击
        pyautogui.rightClick(x, y)
    elif action == "down":
        # 点击下面的坐标
        pyautogui.doubleClick(x, y + 20)


def match_template(image_path, template_path):
    # 读取图像
    image = cv2.imread(image_path)
    template = cv2.imread(template_path)
    # 返回匹配的结果
    result = cv2.matchTemplate(image, template, cv2.TM_CCOEFF_NORMED)
    min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
    top_left = max_loc
    bottom_right = (top_left[0] + template.shape[1], top_left[1] + template.shape[0])

    # 4.画矩形
    cv2.rectangle(image, top_left, bottom_right, (255, 0, 0), 5)
    cv2.imwrite(image_path, image)
    m = mouse.Controller()
    m.position = top_left
    # 返回左上角、右下角的坐标
    return top_left, bottom_right


def get_click_point(top_left, bottom_right):
    # 鼠标左键点击的坐标
    x = (top_left[0] + bottom_right[0]) / 2
    y = (top_left[1] + bottom_right[1]) / 2
    return x, y


"""
初始化路径，没有则创建，兼容末尾有斜杠和没斜杠
"""


def init_path(target_path):
    result_path = treat_path(target_path, True)
    os.makedirs(result_path, exist_ok=True)
    return result_path


"""
处理路径
# param path: 路径信息
# param is_append_string:True追加末尾分隔符False不逐年末位分隔符
# return:处理后的路径
"""


def treat_path(path, is_append_string=False):
    if path.endswith("\\"):
        return path
    else:
        if is_append_string is True:
            return path + os.sep
        else:
            return path


"""
读取EXCEL的内容
"""


def read_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active

    data = []
    # 遍历
    for row in ws.iter_rows(values_only=True):
        if row[0] is None:
            break
        data.append(row)
    return data


"""
输入单元格的值(点击并输入对应的值)
"""


def insert_cell_value(target_file_name, insert_value):
    time.sleep(0.5)
    find_picture_click(origin_image_path, target_image_path, target_file_name, "down")
    pyperclip.copy(insert_value)
    time.sleep(0.5)
    pyautogui.hotkey("ctrl", "v")
    time.sleep(1)


# 查找图片存放的路径
target_image_path = "E:\\NC\\picture"
# 截图路径
origin_image_path = "E:\\NC\\picture\\origin"
# 文件路径
filePathName = "E:\\UMS\\RPA\\广东分公司集中结算垫支协同凭证\\7-集中结算垫资协同凭证科目对照表.xlsx"
# 截图保存格式
file_extension = "jpg"

# 准备工作，选择目录，没有则创建
target_image_path = init_path(target_image_path)
origin_image_path = init_path(origin_image_path)

# -------------解析需要插入的数据------------

# 读取excel的内容
excel_data = read_excel(filePathName)

# # 勾选其中一条集中结算垫支的数据
# find_picture_click(origin_image_path, target_image_path,  "checkbox.jpg", "single")
# # 点击本方修改
# find_picture_click(origin_image_path, target_image_path,  "benfang_modify.jpg", "single")
# time.sleep(5)

# # 点击本方凭证标题
# find_picture_click(origin_image_path, target_image_path, "benfang_voucher_title.jpg", "single")

# time.sleep(0.5)
# for i in range(7):
#     time.sleep(0.5)
#     keyboard.press_and_release("Tab")
# time.sleep(1)
# 
# # 全选复制内容（一条集中结算垫支协同凭证）
# pyautogui.hotkey("ctrl", "a")
# time.sleep(0.5)
# pyautogui.hotkey("ctrl", "c")
# # 获取剪切板的内容
# content = clipboard_get()

# 测试数据
content = """2	1210 集中结算手续费垫支	224108050302\其他应付款\应付业务资金\结算资金\本地业务本金垫款\应付总公司	【客商：200/总部资金结算中心】	CNY	63511161.40	63511161.40		
3	1210 集中结算手续费垫支	224108050502\其他应付款\应付业务资金\结算资金\本地业务手续费垫款\应付总公司	【客商：200/总部资金结算中心】	CNY	62962.68		62962.68	
4	1210 集中结算付款类垫支	224108050302\其他应付款\应付业务资金\结算资金\本地业务本金垫款\应付总公司	【客商：200/总部资金结算中心】	CNY	63651358.11		63651358.11	
5	1210 集中接入付款类垫支	224108050402\其他应付款\应付业务资金\结算资金\接总业务本金垫款\应付总公司	【客商：200/总部资金结算中心】	CNY	16582013.18	16582013.18		
6	1210 集中接入付款类垫支	224108050402\其他应付款\应付业务资金\结算资金\接总业务本金垫款\应付总公司	【客商：200/总部资金结算中心】	CNY	16533168.84		16533168.84	
7	1210 集中接入手续费垫支	224108050602\其他应付款\应付业务资金\结算资金\接总业务手续费垫款\应付总公司	【客商：200/总部资金结算中心】	CNY	72978.17		72978.17	"""

# print('剪切板的内容是：', content)
# 如果这里剪切板的内容为空处理
rowData = []
rowData = content.splitlines()

print(len(excel_data))

# 需要插入的数据(集中结算手续费垫支)
insert_list = []

temp_list = []

# 遍历每一行数据
for i in range(len(excel_data)):
    # print("行的数据是：", row)
    temp_row = []
    for row in rowData:
        # 查找手续费科目
        if excel_data[i][0] in row:
            # print("被选中的值：{},行的数据是：{}", excel_data[i][0], row)
            # 解析row
            columns = row.split("	")
            # 摘要:手续费科目:手续费客商:原币:借方:贷方
            insert_row = [columns[1], excel_data[i][1], excel_data[i][2], columns[5], columns[7], columns[6]]
            # 加入到列表中
            insert_list.append(insert_row)

        # 付款类（接入付款、结算付款）
        elif excel_data[i][5] in row:
            # 列表不为空
            if temp_row:
                # 解析row
                columns = row.split("	")
                # 借方和
                temp_row[3] = Decimal(temp_row[3] if temp_row[3] != '' else 0) + Decimal(
                    columns[6] if columns[6] != '' else 0)
                temp_row[3] = temp_row[3].quantize(Decimal("0.00"))
                # 贷方和
                temp_row[4] = Decimal(temp_row[4] if temp_row[4] != '' else 0) + Decimal(
                    columns[7] if columns[7] != '' else 0)
                temp_row[4] = temp_row[4].quantize(Decimal("0.00"))
            else:
                # 解析row
                columns = row.split("	")
                # 摘要
                temp_row.append(columns[1])
                # 科目
                temp_row.append(excel_data[i][6])
                # 客商
                temp_row.append(excel_data[i][7])
                # # 原币
                # pay_row.append(columns[5])
                # 借方
                temp_row.append(columns[6])
                # 贷方
                temp_row.append(columns[7])
    if temp_row:
        temp_list.append(temp_row)
        print('每一行的数据：', temp_row)

for row in temp_list:
    pay_row = row
    # 原币
    functional_currency_amount = 0
    # 借方
    debit_amount = 0
    # 贷方
    credit_amount = 0
    # 借方和-贷方和
    debet_sum = Decimal(pay_row[3] if pay_row[3] != '' else 0).quantize(Decimal("0.00"))
    credit_sum = Decimal(pay_row[4] if pay_row[4] != '' else 0).quantize(Decimal("0.00"))
    print("两个数的值是：{},{}", debet_sum, credit_sum)
    balance = debet_sum - credit_sum
    print("balance的值是：", balance.quantize(Decimal("0.00")))
    if balance > 0:
        # 贷方
        credit_amount = balance
        functional_currency_amount = balance
    elif balance < 0:
        # 借方
        debit_amount = abs(balance)
        functional_currency_amount = abs(balance)
    # 摘要
    pay_row.append(pay_row[0])
    # 科目
    pay_row.append(pay_row[1])
    # 客商
    pay_row.append(pay_row[2])
    # # 原币
    pay_row.append(functional_currency_amount)
    # 借方
    pay_row.append(debit_amount)
    # 贷方
    pay_row.append(credit_amount)
    insert_list.append(pay_row)
print('insert_list', insert_list)
# -------------解析需要插入的数据------------

# -------------插入单挑凭证的操作------------
# 点击本方凭证标题
find_picture_click(origin_image_path, target_image_path, "benfang_voucher_title.jpg", "single")

time.sleep(0.5)
for i in range(6):
    time.sleep(0.5)
    keyboard.press_and_release("Tab")

time.sleep(1)


for row in insert_list:
    # # 开始单条输入
    time.sleep(1)
    # 点击本方插入
    find_picture_click(origin_image_path, target_image_path, "insert_voucher.jpg", "single")

    # 点击本方凭证标题
    find_picture_click(origin_image_path, target_image_path, "benfang_voucher_title.jpg", "single")

    time.sleep(0.5)
    for i in range(5):
        time.sleep(0.5)
        keyboard.press_and_release("Tab")

    # 摘要
    insert_cell_value('benfang_modify_summary.jpg', row[0])

    # 科目
    insert_cell_value('bengfang_modify_subject.jpg', row[1])

    # 辅助核算
    find_picture_click(origin_image_path, target_image_path, "subsidiary.jpg", "down")
    time.sleep(0.5)
    find_picture_click(origin_image_path, target_image_path, "search_customer.jpg", "single")
    time.sleep(1.5)
    find_picture_click(origin_image_path, target_image_path, "account_content.jpg", "down")
    time.sleep(1)
    pyperclip.copy(row[2])
    pyautogui.hotkey("ctrl", "v")
    time.sleep(2)
    find_picture_click(origin_image_path, target_image_path, "account_confirm.jpg", "single")
    time.sleep(3)

    # 原币
    insert_cell_value('original_currency.jpg', str(row[3]))

    # 借方
    debet_amount = str(row[4])
    if debet_amount:
        insert_cell_value('debet.jpg', debet_amount)

    # 贷方
    credit_amount = str(row[5])
    if credit_amount:
        insert_cell_value('credit.jpg', credit_amount)
# -------------插入单挑凭证的操作------------
import time
import subprocess
import pygetwindow as gw
import cv2
import pyautogui
import keyboard
import os
from pynput import mouse
from PIL import ImageGrab
import pyperclip
from openpyxl import load_workbook
from decimal import Decimal

"""获取剪贴板数据"""


def clipboard_get():
    time.sleep(0.2)
    data = pyperclip.paste()  # 主要这里差别
    return data


"""
# 查找输入框并输入字符串
"""

# def find_picture_click_and_input(orginal_path, target_path, original_name, target_name, input_word):
#     find_picture_click(orginal_path, target_path, original_name, target_name, "double")
#     # 输入字符串
#     pyautogui.typewrite( interval=0.1)
#     pyautogui.typewrite(input_word, interval=0.1)
#     keyboard.press_and_release("enter")


"""
# 查找图片并点击
# action : single: 单击 , double: 双击
"""


def find_picture_click(orginal_path, target_path, filename, action, extra_x=0, extra_y=0):
    # 找到最后一个圆点的索引位置
    last_dot_index = filename.rfind(".")
    if last_dot_index != -1:
        file_name = filename[:last_dot_index]  # 获取文件名部分
        file_extension = filename[last_dot_index + 1:]  # 获取文件名后缀部分
    else:
        print("无法解析文件名和文件名后缀")
    # 截图的保存路径的绝对路径名
    orginal_path_name = orginal_path + file_name + "_origin." + file_extension
    # 目标图片的保存路径的绝对路径名
    target_path_name = target_path + file_name + "." + file_extension

    # 截图
    pic = ImageGrab.grab()

    # 保存图片
    pic.save(orginal_path_name)

    # 获取目标图片左上角、右下角的坐标
    top_left, bottom_right = match_template(orginal_path_name, target_path_name)
    # print("左上角、右下角的坐标分别是：", (top_left, bottom_right))

    # 获取鼠标即将点击的x,y坐标
    x, y = get_click_point(top_left, bottom_right)
    # print("鼠标点击的[x,y]坐标是：", (x, y))

    if action == "single":
        # 单击鼠标左键
        # 参数1：移动x坐标到指定位置
        # 参数2：移动y坐标到指定位置
        # 参数3：点击次数
        # 无参数：在当前坐标单击
        pyautogui.click(x + extra_x, y + extra_y)
    elif action == "double":
        # 双击鼠标左键
        # 参数1：移动x坐标到指定位置
        # 参数2：移动y坐标到指定位置
        # 无参数：在当前坐标双击
        pyautogui.doubleClick(x + extra_x, y + extra_y)
    elif action == "right":
        # 点击鼠标右键
        # 参数1：移动x坐标到指定位置
        # 参数2：移动y坐标到指定位置
        # 无参数：在当前坐标点击
        pyautogui.rightClick(x + extra_x, y + extra_y)


def match_template(image_path, template_path):
    # 读取图像
    image = cv2.imread(image_path)
    template = cv2.imread(template_path)
    # 返回匹配的结果
    result = cv2.matchTemplate(image, template, cv2.TM_CCOEFF_NORMED)
    min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
    top_left = max_loc
    bottom_right = (top_left[0] + template.shape[1], top_left[1] + template.shape[0])

    # 4.画矩形
    cv2.rectangle(image, top_left, bottom_right, (255, 0, 0), 5)
    cv2.imwrite(image_path, image)
    m = mouse.Controller()
    m.position = top_left
    # 返回左上角、右下角的坐标
    return top_left, bottom_right


def get_click_point(top_left, bottom_right):
    # 鼠标左键点击的坐标
    x = (top_left[0] + bottom_right[0]) / 2
    y = (top_left[1] + bottom_right[1]) / 2
    return x, y


"""
初始化路径，没有则创建，兼容末尾有斜杠和没斜杠
"""


def init_path(target_path):
    result_path = treat_path(target_path, True)
    os.makedirs(result_path, exist_ok=True)
    return result_path


"""
处理路径
# param path: 路径信息
# param is_append_string:True追加末尾分隔符False不逐年末位分隔符
# return:处理后的路径
"""


def treat_path(path, is_append_string=False):
    if path.endswith("\\"):
        return path
    else:
        if is_append_string is True:
            return path + os.sep
        else:
            return path


"""
读取EXCEL的内容
"""


def read_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active

    data = []
    # 遍历
    for row in ws.iter_rows(values_only=True):
        if row[0] is None:
            break
        data.append(row)
    return data


"""
输入单元格的值(点击并输入对应的值)
"""


def insert_cell_value(target_file_name, insert_value):
    time.sleep(0.5)
    find_picture_click(origin_image_path, target_image_path, target_file_name, 0, 50)
    pyperclip.copy(insert_value)
    time.sleep(0.5)
    pyautogui.hotkey("ctrl", "v")
    time.sleep(1)


# 查找图片存放的路径
target_image_path = "E:\\NC\\picture"
# 截图路径
origin_image_path = "E:\\NC\\picture\\origin"
# 文件路径
filePathName = "E:\\UMS\\RPA\\广东分公司集中结算垫支协同凭证\\7-集中结算垫资协同凭证科目对照表.xlsx"
# 截图保存格式
file_extension = "jpg"

# 准备工作，选择目录，没有则创建
target_image_path = init_path(target_image_path)
origin_image_path = init_path(origin_image_path)

# -------------解析需要插入的数据------------

# 读取excel的内容
excel_data = read_excel(filePathName)

# 勾选其中一条集中结算垫支的数据
find_picture_click(origin_image_path, target_image_path,  "checkbox_template.jpg", "single")
# 点击本方修改
find_picture_click(origin_image_path, target_image_path,  "benfang_modify.jpg", "single")
time.sleep(5)

# 点击本方凭证标题
find_picture_click(origin_image_path, target_image_path, "benfang_voucher_title.jpg", "single")

time.sleep(0.5)
for i in range(7):
    time.sleep(0.5)
    keyboard.press_and_release("Tab")
time.sleep(1)

# 全选复制内容（一条集中结算垫支协同凭证）
pyautogui.hotkey("ctrl", "a")
time.sleep(0.5)
pyautogui.hotkey("ctrl", "c")
# 获取剪切板的内容
content = clipboard_get()

# 测试数据
content = """2	1210 集中结算手续费垫支	224108050302\其他应付款\应付业务资金\结算资金\本地业务本金垫款\应付总公司	【客商：200/总部资金结算中心】	CNY	63511161.40	63511161.40		
3	1210 集中结算手续费垫支	224108050502\其他应付款\应付业务资金\结算资金\本地业务手续费垫款\应付总公司	【客商：200/总部资金结算中心】	CNY	62962.68		62962.68	
4	1210 集中结算付款类垫支	224108050302\其他应付款\应付业务资金\结算资金\本地业务本金垫款\应付总公司	【客商：200/总部资金结算中心】	CNY	63651358.11		63651358.11	
5	1210 集中接入付款类垫支	224108050402\其他应付款\应付业务资金\结算资金\接总业务本金垫款\应付总公司	【客商：200/总部资金结算中心】	CNY	16582013.18	16582013.18		
6	1210 集中接入付款类垫支	224108050402\其他应付款\应付业务资金\结算资金\接总业务本金垫款\应付总公司	【客商：200/总部资金结算中心】	CNY	16533168.84		16533168.84	
7	1210 集中接入手续费垫支	224108050602\其他应付款\应付业务资金\结算资金\接总业务手续费垫款\应付总公司	【客商：200/总部资金结算中心】	CNY	72978.17		72978.17	"""

# print('剪切板的内容是：', content)
# 如果这里剪切板的内容为空处理
rowData = []
rowData = content.splitlines()

print(len(excel_data))

# 需要插入的数据(集中结算手续费垫支)
insert_list = []

temp_list = []

# 遍历每一行数据
for i in range(len(excel_data)):
    # print("行的数据是：", row)
    temp_row = []
    for row in rowData:
        # 查找手续费科目
        if excel_data[i][0] in row:
            # print("被选中的值：{},行的数据是：{}", excel_data[i][0], row)
            # 解析row
            columns = row.split("	")
            # 摘要:手续费科目:手续费客商:原币:借方:贷方
            insert_row = [columns[1], excel_data[i][1], excel_data[i][2], columns[5], columns[7], columns[6]]
            # 加入到列表中
            insert_list.append(insert_row)

        # 付款类（接入付款、结算付款）
        elif excel_data[i][5] in row:
            # 列表不为空
            if temp_row:
                # 解析row
                columns = row.split("	")
                # 借方和
                temp_row[3] = Decimal(temp_row[3] if temp_row[3] != '' else 0) + Decimal(
                    columns[6] if columns[6] != '' else 0)
                temp_row[3] = temp_row[3].quantize(Decimal("0.00"))
                # 贷方和
                temp_row[4] = Decimal(temp_row[4] if temp_row[4] != '' else 0) + Decimal(
                    columns[7] if columns[7] != '' else 0)
                temp_row[4] = temp_row[4].quantize(Decimal("0.00"))
            else:
                # 解析row
                columns = row.split("	")
                # 摘要
                temp_row.append(columns[1])
                # 科目
                temp_row.append(excel_data[i][6])
                # 客商
                temp_row.append(excel_data[i][7])
                # # 原币
                # pay_row.append(columns[5])
                # 借方
                temp_row.append(columns[6])
                # 贷方
                temp_row.append(columns[7])
    if temp_row:
        temp_list.append(temp_row)
        print('每一行的数据：', temp_row)

for row in temp_list:
    pay_row = row
    # 原币
    functional_currency_amount = 0
    # 借方
    debit_amount = 0
    # 贷方
    credit_amount = 0
    # 借方和-贷方和
    debet_sum = Decimal(pay_row[3] if pay_row[3] != '' else 0).quantize(Decimal("0.00"))
    credit_sum = Decimal(pay_row[4] if pay_row[4] != '' else 0).quantize(Decimal("0.00"))
    print("两个数的值是：{},{}", debet_sum, credit_sum)
    balance = debet_sum - credit_sum
    print("balance的值是：", balance.quantize(Decimal("0.00")))
    if balance > 0:
        # 贷方
        credit_amount = balance
        functional_currency_amount = balance
    elif balance < 0:
        # 借方
        debit_amount = abs(balance)
        functional_currency_amount = abs(balance)
    # 摘要
    pay_row.append(pay_row[0])
    # 科目
    pay_row.append(pay_row[1])
    # 客商
    pay_row.append(pay_row[2])
    # # 原币
    pay_row.append(functional_currency_amount)
    # 借方
    pay_row.append(debit_amount)
    # 贷方
    pay_row.append(credit_amount)
    insert_list.append(pay_row)
print('insert_list', insert_list)
# -------------解析需要插入的数据------------

# -------------插入单挑凭证的操作------------
# 点击本方凭证标题
find_picture_click(origin_image_path, target_image_path, "benfang_voucher_title.jpg", "single")

time.sleep(0.5)
for i in range(6):
    time.sleep(0.5)
    keyboard.press_and_release("Tab")

time.sleep(1)


for row in insert_list:
    # # 开始单条输入
    time.sleep(1)
    # 点击本方插入
    find_picture_click(origin_image_path, target_image_path, "insert_voucher.jpg", "single")

    # 点击本方凭证标题
    find_picture_click(origin_image_path, target_image_path, "benfang_voucher_title.jpg", "single")

    time.sleep(0.5)
    for i in range(5):
        time.sleep(0.5)
        keyboard.press_and_release("Tab")

    # 摘要
    insert_cell_value('benfang_modify_summary.jpg', row[0])

    # 科目
    insert_cell_value('bengfang_modify_subject.jpg', row[1])

    # 辅助核算
    find_picture_click(origin_image_path, target_image_path, "subsidiary.jpg",0,20)
    time.sleep(0.5)
    find_picture_click(origin_image_path, target_image_path, "search_customer.jpg", "single")
    time.sleep(1.5)
    find_picture_click(origin_image_path, target_image_path, "account_content.jpg", 0,20)
    time.sleep(1)
    pyperclip.copy(row[2])
    pyautogui.hotkey("ctrl", "v")
    time.sleep(2)
    find_picture_click(origin_image_path, target_image_path, "account_confirm.jpg", "single")
    time.sleep(3)

    # 原币
    insert_cell_value('original_currency.jpg', str(row[3]))

    # 借方
    debet_amount = str(row[4])
    if debet_amount:
        insert_cell_value('debet.jpg', debet_amount)

    # 贷方
    credit_amount = str(row[5])
    if credit_amount:
        insert_cell_value('credit.jpg', credit_amount)

def process_get_voucher_info(original_path, target_path, file_name):
    # 凭证号
    # find_picture_click(origin_image_path, target_image_path, file_name, "double", -50, 0)
    # pyautogui.hotkey("ctrl", "c")
    # voucher = clipboard_get()
    # print("获取到的凭证号是：", voucher)

    # 会计日期 先截图
    # result = handle_path_file_name(original_path, target_path, file_name)
    result = {}
    original_path_name, target_path_name = result

    # 获取目标图片左上角、右下角的坐标
    top_left, bottom_right = match_template(original_path_name, target_path_name)
    start_x = bottom_right[0]
    start_y = bottom_right[1]
    end_x = bottom_right[0] + 80
    end_y = bottom_right[1] + 24
    print(top_left)
    print(bottom_right)
    start_x = top_left[0]
    start_y = top_left[1]
    end_x = bottom_right[0]
    end_y = bottom_right[1]

    # # 打开原始图像
    # image = Image.open(original_path_name)
    # # 创建绘图对象
    # draw = ImageDraw.Draw(image)
    # # 绘制矩形边框
    # draw.rectangle([(start_x, start_y), (end_x, end_y)], outline='red', width=3)

    # # 截取指定区域
    cropped_image = image.crop((start_x, start_y, end_x, end_y))
    # 保存截取后的图像
    cropped_image.save(original_path + 'account_period_date.jpg')

# -------------插入单挑凭证的操作------------


# --------------------------------------------------

# 切换业务日期
# date = excel_data[1][10].strftime('%Y-%m-%d')
# find_picture_click(origin_image_path, target_image_path, "ask.jpg", "single", -50)
# time.sleep(1)
# find_picture_click(origin_image_path, target_image_path, "business_date.jpg", "double", 50)
# # 复制日期，粘贴
# pyperclip.copy(date)
# time.sleep(0.5)
# pyautogui.hotkey("ctrl", "v")
# 
# # 点击确定
# find_picture_click(origin_image_path, target_image_path, "business_date_confirm.jpg", "single")
