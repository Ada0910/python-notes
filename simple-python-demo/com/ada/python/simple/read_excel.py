"""
Python读取excel(方式一)
$ pip3 install pandas

"""
# 1.导入pandas模块
# import pandas as pd
# 
# # 2.把Excel文件中的数据读入pandas
# df = pd.read_excel('Python招聘数据（全）.xlsx')
# print(df)
# # 3.读取excel的某一个sheet
# df = pd.read_excel('Python招聘数据（全）.xlsx', sheet_name='Sheet1')
# print(df)
# # 4.获取列标题
# print(df.columns)
# # 5.获取列行标题
# print(df.index)
# # 6.制定打印某一列
# print(df["工资水平"])
# # 7.描述数据
# print(df.describe())


"""
Python读取excel(方式二)
$ pip3 install openpyxl
"""
from openpyxl import load_workbook
# 1.打开 Excel 表格并获取表格名称
workbook = load_workbook(filename="E:\\UMS\\RPA\\广东分公司集中结算垫支协同凭证\\7-集中结算垫资协同凭证科目对照表.xlsx")
print(workbook.sheetnames)
# 2.通过 sheet 名称获取表格
sheet = workbook["Sheet1"]
print(sheet)
# 3.获取表格的尺寸大小(几行几列数据) 这里所说的尺寸大小，指的是 excel 表格中的数据有几行几列，针对的是不同的 sheet 而言。
print(sheet.dimensions)
cell = sheet[sheet.dimensions]
print(cell)
for i in cell:
    for j in i:
        print(j.value)
# 4.获取表格内某个格子的数据
# 1 sheet["A1"]方式
cell1 = sheet["A1"]
cell2 = sheet["C11"]
print(cell1.value, cell2.value)
"""
workbook.active 打开激活的表格; sheet["A1"] 获取 A1 格子的数据; cell.value 获取格子中的值;
"""
# 4.2sheet.cell(row=, column=)方式
cell1 = sheet.cell(row = 1,column = 1)
cell2 = sheet.cell(row = 11,column = 3)
print(cell1.value, cell2.value)

# 5. 获取一系列格子
# 获取 A1:C2 区域的值
cell = sheet["A1:C2"]
print(cell)
for i in cell:
    for j in i:
        print(j.value)

        
"""
方式三：只支持excel的xls格式
$ pip3 install xlrd xlwt xlutils

"""
# 导入 xlrd 库
# import xlrd
# # 打开刚才我们写入的 test_w.xls 文件
# wb = xlrd.open_workbook("Python招聘数据（全）.xlsx")
# # 获取并打印 sheet 数量
# print( "sheet 数量:", wb.nsheets)
# # 获取并打印 sheet 名称
# print( "sheet 名称:", wb.sheet_names())
# # 根据 sheet 索引获取内容
# sh1 = wb.sheet_by_index(0)
# # 也可根据 sheet 名称获取内容
# # sh = wb.sheet_by_name('成绩')
# # 获取并打印该 sheet 行数和列数
# print( u"sheet %s 共 %d 行 %d 列" % (sh1.name, sh1.nrows, sh1.ncols))
# # 获取并打印某个单元格的值
# print( "第一行第二列的值为:", sh1.cell_value(0, 1))
# # 获取整行或整列的值
# rows = sh1.row_values(0) # 获取第一行内容
# cols = sh1.col_values(1) # 获取第二列内容
# # 打印获取的行列值
# print( "第一行的值为:", rows)
# print( "第二列的值为:", cols)
# # 获取单元格内容的数据类型
# print( "第二行第一列的值类型为:", sh1.cell(1, 0).ctype)