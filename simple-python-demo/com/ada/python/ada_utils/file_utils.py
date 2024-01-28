"""
文件处理工具类
"""
from openpyxl import load_workbook

"""
解析excel，并返回数据

Args:
    filename：文件名，全路径
Returns:
    返回：二维数组

Raises:
    异常情况的说明（如果有的话）

"""


def read_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active

    data = []
    # 遍历
    for row in ws.iter_rows(values_only=True):
        if row[0] is None:
            break
        data.append(row)
    return data